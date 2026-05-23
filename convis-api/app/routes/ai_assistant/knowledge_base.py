"""
Knowledge Base endpoints for AI Assistants
"""
import re
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from app.models.ai_assistant import FileUploadResponse, KnowledgeBaseFile, DeleteResponse
from app.config.database import Database
from app.utils import conversational_rag, mongo_rag
from app.utils.assistant_keys import resolve_assistant_api_key
from app.utils.auth import get_current_user, verify_user_ownership
from bson import ObjectId
from datetime import datetime
import os
import shutil
import json
import logging

logger = logging.getLogger(__name__)

# Router-level auth: every endpoint here requires a valid JWT. Per-route
# `verify_user_ownership` further checks the caller owns the assistant.
router = APIRouter(dependencies=[Depends(get_current_user)])


def _safe_filename_for_disk(raw: str) -> str:
    """Strip path separators and any character that could traverse / overwrite
    arbitrary files. The Mongo `filename` field still stores the user's
    original name (used for display + lookup); only the on-disk filename is
    sanitized.
    """
    base = os.path.basename(raw or "").replace("\x00", "")
    cleaned = re.sub(r"[^A-Za-z0-9._-]", "_", base)[:200]
    return cleaned or "upload"

# Base directory for uploads
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "../../../uploads/knowledge_base")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {'.pdf', '.docx', '.doc', '.xlsx', '.xls', '.txt'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


def get_file_extension(filename: str) -> str:
    """Get file extension from filename"""
    return os.path.splitext(filename)[1].lower()


@router.post("/{assistant_id}/upload", response_model=FileUploadResponse, status_code=status.HTTP_200_OK)
async def upload_knowledge_base_file(
    assistant_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Upload a knowledge base file for an AI assistant

    Args:
        assistant_id: AI Assistant ID
        file: Uploaded file (PDF, DOCX, XLSX, TXT)

    Returns:
        FileUploadResponse: Upload status and file info
    """
    try:
        db = Database.get_db()
        assistants_collection = db['assistants']

        logger.info(f"Uploading file {file.filename} for assistant {assistant_id}")

        # Validate assistant_id
        try:
            assistant_obj_id = ObjectId(assistant_id)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid assistant_id format"
            )

        # Fetch assistant
        assistant = assistants_collection.find_one({"_id": assistant_obj_id})
        if not assistant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="AI assistant not found"
            )
        await verify_user_ownership(current_user, str(assistant.get('user_id')))

        # Validate file extension
        file_ext = get_file_extension(file.filename)
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File type not supported. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
            )

        # Read file content
        content = await file.read()
        file_size = len(content)

        # Check file size
        if file_size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)}MB"
            )

        # Create assistant-specific directory
        assistant_dir = os.path.join(UPLOAD_DIR, str(assistant_id))
        os.makedirs(assistant_dir, exist_ok=True)

        # Generate unique filename. We sanitize the user-supplied name BEFORE
        # using it in any filesystem path so a crafted filename like
        # "../../../etc/passwd" can't escape the per-assistant upload dir.
        # The Mongo `filename` field still uses file.filename verbatim so the
        # dashboard shows the original name.
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        safe_filename = f"{timestamp}_{_safe_filename_for_disk(file.filename)}"
        file_path = os.path.join(assistant_dir, safe_filename)
        # Defense in depth: ensure the resolved path stays inside assistant_dir.
        if not os.path.realpath(file_path).startswith(
            os.path.realpath(assistant_dir) + os.sep
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid filename",
            )

        # Save file
        with open(file_path, 'wb') as f:
            f.write(content)

        logger.info(f"File saved to {file_path}")

        # Get OpenAI API key for processing
        try:
            openai_api_key, _ = resolve_assistant_api_key(db, assistant, required_provider="openai")
        except HTTPException as exc:
            os.remove(file_path)
            raise exc

        # Extract text once (PDF/DOCX/XLSX/TXT — implementations live in
        # conversational_rag and are file-format-correct). Then index into
        # MongoDB via mongo_rag — the agent reads from there during calls.
        # We deliberately do NOT use conversational_rag's ChromaDB path: it
        # stores chunks on the API task's ephemeral filesystem which the
        # agent's separate Fargate task can't read.
        logger.info(f"Extracting text from {file.filename}...")
        text = conversational_rag.extract_text_from_file(file_path, file_ext.replace('.', ''))
        if not text or not text.strip():
            os.remove(file_path)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not extract text from this file (possibly an image-only PDF)."
            )

        logger.info(f"Indexing {len(text)} chars into MongoDB knowledge_chunks...")
        result = mongo_rag.index_document(
            assistant_id=assistant_id,
            text=text,
            filename=file.filename,
            file_type=file_ext.replace('.', ''),
            api_key=openai_api_key,
        )

        if not result['success']:
            os.remove(file_path)
            error_msg = result.get('error', 'Unknown error')

            # Provide helpful error messages for common issues
            if 'OCR not available' in error_msg or 'Tesseract' in error_msg:
                error_msg = (
                    "This PDF appears to be image-based/scanned and requires OCR to extract text. "
                    "Please install Tesseract OCR:\n"
                    "Ubuntu/Debian: sudo apt-get install tesseract-ocr tesseract-ocr-eng\n"
                    "macOS: brew install tesseract\n"
                    "Or convert your PDF to a text-searchable PDF first."
                )
            elif 'Could not extract text' in error_msg:
                error_msg = (
                    "Failed to extract text from PDF. This may be an image-based/scanned document. "
                    "Try converting it to a text-searchable PDF or install Tesseract OCR for automatic text recognition."
                )

            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=error_msg
            )

        # Store file metadata
        file_metadata = {
            "filename": file.filename,
            "file_type": file_ext.replace('.', ''),
            "file_size": file_size,
            "uploaded_at": datetime.utcnow(),
            "file_path": file_path,
            "chunks_count": result['chunks_count']
        }

        # Update assistant document
        assistants_collection.update_one(
            {"_id": assistant_obj_id},
            {
                "$push": {"knowledge_base_files": file_metadata},
                "$set": {"updated_at": datetime.utcnow()}
            }
        )

        logger.info(f"Successfully processed {result['chunks_count']} chunks from {file.filename}")

        # Get total files count
        updated_assistant = assistants_collection.find_one({"_id": assistant_obj_id})
        total_files = len(updated_assistant.get('knowledge_base_files', []))

        return FileUploadResponse(
            message=f"File uploaded and processed successfully. Created {result['chunks_count']} knowledge chunks.",
            file=KnowledgeBaseFile(
                filename=file.filename,
                file_type=file_ext.replace('.', ''),
                file_size=file_size,
                uploaded_at=file_metadata['uploaded_at'].isoformat() + "Z",
                file_path=file_path
            ),
            total_files=total_files
        )

    except HTTPException:
        raise
    except Exception as error:
        import traceback
        logger.error(f"Error uploading knowledge base file: {str(error)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload file: {str(error)}"
        )


@router.get("/{assistant_id}/preview/{filename}", status_code=status.HTTP_200_OK)
async def preview_document_content(
    assistant_id: str,
    filename: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Get the extracted text content of a knowledge base document

    Args:
        assistant_id: AI Assistant ID
        filename: Name of the file to preview

    Returns:
        JSON with extracted_text field
    """
    try:
        db = Database.get_db()
        assistants_collection = db['assistants']

        logger.info(f"Fetching content for file {filename} from assistant {assistant_id}")

        # Validate assistant_id
        try:
            assistant_obj_id = ObjectId(assistant_id)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid assistant_id format"
            )

        # Fetch assistant
        assistant = assistants_collection.find_one({"_id": assistant_obj_id})
        if not assistant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="AI assistant not found"
            )
        await verify_user_ownership(current_user, str(assistant.get('user_id')))

        # Find file in knowledge base
        kb_files = assistant.get('knowledge_base_files', [])
        file_info = None
        for f in kb_files:
            if f['filename'] == filename:
                file_info = f
                break

        if not file_info:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in knowledge base"
            )

        # Read the file and extract text
        file_path = file_info['file_path']
        if not os.path.exists(file_path):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Physical file not found"
            )

        # Extract text based on file type
        file_ext = get_file_extension(filename)
        extracted_text = ""

        try:
            if file_ext == '.pdf':
                import PyPDF2
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    num_pages = len(pdf_reader.pages)

                    # Try to extract text from each page
                    for page_num, page in enumerate(pdf_reader.pages, 1):
                        page_text = page.extract_text()
                        if page_text and page_text.strip():
                            extracted_text += f"--- Page {page_num} ---\n{page_text}\n\n"

                    # If no text was extracted, this might be a scanned/image-based PDF
                    if not extracted_text.strip():
                        extracted_text = f"""📄 PDF Document Information:

Filename: {filename}
Total Pages: {num_pages}
File Size: {file_info['file_size'] / 1024:.2f} KB

⚠️ Note: This PDF appears to be image-based or scanned. No text could be extracted directly.

The document has been processed and stored in the knowledge base. During conversations, the AI assistant will be able to access the embedded content from this document.

To view the actual content, please:
1. Open the PDF file directly
2. Or convert it to a text-searchable PDF using OCR software

The AI can still reference information from this document during conversations, as it was processed during upload."""

            elif file_ext in ['.docx', '.doc']:
                from docx import Document
                doc = Document(file_path)
                for paragraph in doc.paragraphs:
                    if paragraph.text.strip():
                        extracted_text += paragraph.text + "\n"

                if not extracted_text.strip():
                    extracted_text = f"""📄 Word Document Information:

Filename: {filename}
File Size: {file_info['file_size'] / 1024:.2f} KB

⚠️ Note: This document appears to be empty or contains only non-text elements (images, tables, etc.).

The document has been processed and stored in the knowledge base."""

            elif file_ext in ['.xlsx', '.xls']:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    extracted_text += f"--- Sheet: {sheet_name} ---\n"
                    row_count = 0
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                            extracted_text += row_text + "\n"
                            row_count += 1
                    extracted_text += f"\n({row_count} rows)\n\n"

                if not extracted_text.strip():
                    extracted_text = f"""📊 Excel Document Information:

Filename: {filename}
File Size: {file_info['file_size'] / 1024:.2f} KB

⚠️ Note: This spreadsheet appears to be empty.

The document has been processed and stored in the knowledge base."""

            elif file_ext == '.txt':
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                    extracted_text = file.read()

                if not extracted_text.strip():
                    extracted_text = f"""📝 Text Document Information:

Filename: {filename}
File Size: {file_info['file_size'] / 1024:.2f} KB

⚠️ Note: This text file appears to be empty.

The document has been processed and stored in the knowledge base."""

            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Unsupported file type"
                )

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error extracting text from {filename}: {str(e)}")
            # Instead of raising an error, return helpful information
            extracted_text = f"""📄 Document Information:

Filename: {filename}
File Type: {file_ext.replace('.', '').upper()}
File Size: {file_info['file_size'] / 1024:.2f} KB

⚠️ Error extracting text content: {str(e)}

The document has been uploaded and stored in the knowledge base. The AI assistant may still be able to access embedded information from this document during conversations.

If you need to view the content, please try:
1. Opening the file directly
2. Converting it to a different format
3. Re-uploading the file"""

        # Ensure we always return something useful
        final_text = extracted_text.strip() if extracted_text and extracted_text.strip() else f"""📄 Document Information:

Filename: {filename}
File Type: {file_ext.replace('.', '').upper()}
File Size: {file_info['file_size'] / 1024:.2f} KB

⚠️ No text content could be extracted from this document.

This may happen if:
- The document is image-based or scanned
- The document is empty
- The document format is not fully supported

The document has been stored in the knowledge base and the AI assistant may still be able to reference it during conversations."""

        return {
            "filename": filename,
            "file_type": file_ext.replace('.', ''),
            "file_size": file_info['file_size'],
            "extracted_text": final_text
        }

    except HTTPException:
        raise
    except Exception as error:
        import traceback
        logger.error(f"Error previewing document: {str(error)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to preview document: {str(error)}"
        )


@router.delete("/{assistant_id}/files/{filename}", response_model=DeleteResponse, status_code=status.HTTP_200_OK)
async def delete_knowledge_base_file(
    assistant_id: str,
    filename: str,
    current_user: dict = Depends(get_current_user),
):
    """
    Delete a knowledge base file from an AI assistant

    Args:
        assistant_id: AI Assistant ID
        filename: Name of the file to delete

    Returns:
        DeleteResponse: Deletion status
    """
    try:
        db = Database.get_db()
        assistants_collection = db['assistants']

        logger.info(f"Deleting file {filename} from assistant {assistant_id}")

        # Validate assistant_id
        try:
            assistant_obj_id = ObjectId(assistant_id)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid assistant_id format"
            )

        # Fetch assistant
        assistant = assistants_collection.find_one({"_id": assistant_obj_id})
        if not assistant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="AI assistant not found"
            )
        await verify_user_ownership(current_user, str(assistant.get('user_id')))

        # Find file in knowledge base
        kb_files = assistant.get('knowledge_base_files', [])
        file_to_delete = None
        for f in kb_files:
            if f['filename'] == filename:
                file_to_delete = f
                break

        if not file_to_delete:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in knowledge base"
            )

        # Delete from MongoDB knowledge_chunks (the live RAG store).
        # The legacy ChromaDB delete is best-effort; we still call it in case
        # any old chunks linger on local disk during the migration.
        deleted = mongo_rag.delete_document(assistant_id, filename)
        logger.info(f"[RAG] Deleted {deleted} chunks for {filename} from MongoDB")
        try:
            conversational_rag.delete_document_from_kb(assistant_id, filename)
        except Exception:
            logger.debug("Legacy ChromaDB delete skipped/failed (expected post-migration)", exc_info=True)

        # Delete physical file
        file_path = file_to_delete['file_path']
        if os.path.exists(file_path):
            os.remove(file_path)

        # Remove file metadata from database
        assistants_collection.update_one(
            {"_id": assistant_obj_id},
            {
                "$pull": {"knowledge_base_files": {"filename": filename}},
                "$set": {"updated_at": datetime.utcnow()}
            }
        )

        logger.info(f"Successfully deleted file {filename}")

        return DeleteResponse(message=f"File '{filename}' deleted successfully")

    except HTTPException:
        raise
    except Exception as error:
        import traceback
        logger.error(f"Error deleting knowledge base file: {str(error)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete file: {str(error)}"
        )
