"""
RAG (Retrieval-Augmented Generation) utilities for knowledge base
"""
import os
import json
from typing import List, Dict, Any
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import logging
from app.utils.local_embeddings import create_embeddings_auto

logger = logging.getLogger(__name__)


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text content from PDF file"""
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text content from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        return ""


def extract_text_from_excel(file_path: str) -> str:
    """Extract text content from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text = ""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n"

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"

        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from Excel: {e}")
        return ""


def extract_text_from_txt(file_path: str) -> str:
    """Extract text content from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception as e:
        logger.error(f"Error reading TXT file: {e}")
        return ""


def extract_text_from_file(file_path: str, file_type: str) -> str:
    """Extract text from various file types"""
    file_type = file_type.lower()

    if file_type == 'pdf' or file_path.endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_type == 'docx' or file_path.endswith('.docx'):
        return extract_text_from_docx(file_path)
    elif file_type in ['xlsx', 'xls'] or file_path.endswith(('.xlsx', '.xls')):
        return extract_text_from_excel(file_path)
    elif file_type == 'txt' or file_path.endswith('.txt'):
        return extract_text_from_txt(file_path)
    else:
        logger.warning(f"Unsupported file type: {file_type}")
        return ""


def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
    """Split text into overlapping chunks for better context"""
    if not text:
        return []

    chunks = []
    start = 0
    text_length = len(text)

    while start < text_length:
        end = start + chunk_size
        chunk = text[start:end]

        # Try to end at a sentence boundary
        if end < text_length:
            last_period = chunk.rfind('.')
            last_newline = chunk.rfind('\n')
            boundary = max(last_period, last_newline)

            if boundary > chunk_size // 2:  # Only use boundary if it's not too early
                end = start + boundary + 1
                chunk = text[start:end]

        chunks.append(chunk.strip())
        start = end - overlap if end < text_length else text_length

    return chunks


def create_embeddings(texts: List[str], api_key: str = None) -> List[List[float]]:
    """Create embeddings using local model or OpenAI (based on EMBEDDING_PROVIDER env var)"""
    try:
        return create_embeddings_auto(texts, api_key=api_key)
    except Exception as e:
        logger.error(f"Error creating embeddings: {e}")
        return []


def search_knowledge_base(
    query: str,
    knowledge_base: List[Dict[str, Any]],
    api_key: str,
    top_k: int = 3
) -> List[Dict[str, Any]]:
    """
    Search knowledge base for relevant chunks

    Args:
        query: User's question
        knowledge_base: List of dicts with 'text' and 'embedding' keys
        api_key: OpenAI API key for creating query embedding
        top_k: Number of top results to return

    Returns:
        List of relevant text chunks with similarity scores
    """
    try:
        if not knowledge_base:
            return []

        # Create embedding for the query
        query_embeddings = create_embeddings_auto([query], api_key=api_key)
        if not query_embeddings:
            return []
        query_embedding = query_embeddings[0]

        # Calculate cosine similarity with all chunks
        results = []
        for item in knowledge_base:
            if 'embedding' not in item or 'text' not in item:
                continue

            similarity = cosine_similarity(
                [query_embedding],
                [item['embedding']]
            )[0][0]

            results.append({
                'text': item['text'],
                'similarity': float(similarity),
                'filename': item.get('filename', 'unknown')
            })

        # Sort by similarity and return top k
        results.sort(key=lambda x: x['similarity'], reverse=True)
        return results[:top_k]

    except Exception as e:
        logger.error(f"Error searching knowledge base: {e}")
        return []


def build_knowledge_base_context(search_results: List[Dict[str, Any]]) -> str:
    """Build context string from search results"""
    if not search_results:
        return ""

    context_parts = ["### Knowledge Base Information:"]
    for i, result in enumerate(search_results, 1):
        context_parts.append(f"\n[Source {i}: {result['filename']} (relevance: {result['similarity']:.2%})]")
        context_parts.append(result['text'])

    return "\n".join(context_parts)


def process_document_for_knowledge_base(
    file_path: str,
    filename: str,
    file_type: str,
    api_key: str
) -> Dict[str, Any]:
    """
    Process a document and prepare it for knowledge base

    Returns:
        Dict with processed chunks and embeddings
    """
    try:
        # Extract text
        text = extract_text_from_file(file_path, file_type)
        if not text:
            raise ValueError("Could not extract text from file")

        # Split into chunks
        chunks = chunk_text(text)
        if not chunks:
            raise ValueError("No chunks created from text")

        logger.info(f"Created {len(chunks)} chunks from {filename}")

        # Create embeddings
        embeddings = create_embeddings(chunks, api_key)
        if not embeddings:
            raise ValueError("Failed to create embeddings")

        # Prepare knowledge base entries
        kb_entries = []
        for chunk, embedding in zip(chunks, embeddings):
            kb_entries.append({
                'text': chunk,
                'embedding': embedding,
                'filename': filename
            })

        return {
            'success': True,
            'chunks_count': len(chunks),
            'kb_entries': kb_entries,
            'text_length': len(text)
        }

    except Exception as e:
        logger.error(f"Error processing document: {e}")
        return {
            'success': False,
            'error': str(e)
        }
