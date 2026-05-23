"""
Conversational RAG - Optimized for real-time voice conversations
Uses ChromaDB for fast vector search and conversation-aware chunking

IMPORTANT: ChromaDB client is initialized LAZILY to avoid blocking Cloud Run startup.
The client is only created when first needed (not at import time).
"""
import os
import re
from typing import List, Dict, Any, Optional
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
import logging
import threading
from app.utils.local_embeddings import create_embeddings_auto

logger = logging.getLogger(__name__)

# ChromaDB client (lazy initialization to avoid blocking Cloud Run startup)
CHROMA_DB_PATH = os.path.join(os.path.dirname(__file__), "../../chroma_db")
_chroma_client = None
_chroma_lock = threading.Lock()


def get_chroma_client():
    """
    Get ChromaDB client with lazy initialization.
    Thread-safe singleton pattern to avoid blocking app startup.
    """
    global _chroma_client

    if _chroma_client is not None:
        return _chroma_client

    with _chroma_lock:
        # Double-check after acquiring lock
        if _chroma_client is not None:
            return _chroma_client

        try:
            logger.info("[CHROMA] Initializing ChromaDB client (lazy load)...")
            import chromadb
            from chromadb.config import Settings

            _chroma_client = chromadb.Client(Settings(
                persist_directory=CHROMA_DB_PATH,
                anonymized_telemetry=False
            ))
            logger.info("[CHROMA] ChromaDB client initialized successfully")
            return _chroma_client
        except Exception as e:
            logger.error(f"[CHROMA] Failed to initialize ChromaDB: {e}")
            raise


def extract_text_from_pdf(file_path: str) -> str:
    """
    Extract text content from PDF file using multiple methods.
    Tries PyMuPDF first, then PyPDF2, and finally OCR for image-based PDFs.
    """
    text = ""

    # Method 1: Try PyMuPDF (fitz) - works better for most PDFs
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()

        # If we got substantial text, return it
        if len(text.strip()) > 100:
            logger.info(f"Extracted {len(text)} characters using PyMuPDF")
            return text.strip()
    except Exception as e:
        logger.warning(f"PyMuPDF extraction failed: {e}")

    # Method 2: Try PyPDF2 as fallback
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"

        # If we got substantial text, return it
        if len(text.strip()) > 100:
            logger.info(f"Extracted {len(text)} characters using PyPDF2")
            return text.strip()
    except Exception as e:
        logger.warning(f"PyPDF2 extraction failed: {e}")

    # Method 3: Try OCR for image-based/scanned PDFs
    try:
        import fitz  # PyMuPDF
        from PIL import Image
        import pytesseract
        import io

        logger.info("Attempting OCR extraction for image-based PDF...")
        doc = fitz.open(file_path)
        ocr_text = ""

        for page_num in range(len(doc)):
            page = doc[page_num]

            # Convert page to image
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better OCR
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))

            # Perform OCR
            try:
                page_text = pytesseract.image_to_string(img)
                ocr_text += page_text + "\n"
                logger.info(f"OCR page {page_num + 1}/{len(doc)}: extracted {len(page_text)} chars")
            except pytesseract.TesseractNotFoundError:
                logger.error("Tesseract not installed. Install with: sudo apt-get install tesseract-ocr")
                doc.close()
                raise ValueError("OCR not available. Please install Tesseract OCR on your system.")
            except Exception as ocr_error:
                logger.warning(f"OCR failed for page {page_num + 1}: {ocr_error}")

        doc.close()

        if len(ocr_text.strip()) > 50:
            logger.info(f"Extracted {len(ocr_text)} characters using OCR")
            return ocr_text.strip()
        else:
            raise ValueError("No text could be extracted from PDF")

    except Exception as e:
        logger.error(f"OCR extraction failed: {e}")
        raise ValueError(f"Could not extract text from PDF. The file may be corrupted or empty. Error: {str(e)}")

    return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text content from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        return ""


def extract_text_from_excel(file_path: str) -> str:
    """Extract text content from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text = ""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== {sheet_name} ===\n"

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"

        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from Excel: {e}")
        return ""


def extract_text_from_txt(file_path: str) -> str:
    """Extract text content from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception as e:
        logger.error(f"Error reading TXT file: {e}")
        return ""


def extract_text_from_file(file_path: str, file_type: str) -> str:
    """Extract text from various file types"""
    file_type = file_type.lower()

    if file_type == 'pdf' or file_path.endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_type == 'docx' or file_path.endswith('.docx'):
        return extract_text_from_docx(file_path)
    elif file_type in ['xlsx', 'xls'] or file_path.endswith(('.xlsx', '.xls')):
        return extract_text_from_excel(file_path)
    elif file_type == 'txt' or file_path.endswith('.txt'):
        return extract_text_from_txt(file_path)
    else:
        logger.warning(f"Unsupported file type: {file_type}")
        return ""


def chunk_text_for_conversation(text: str, chunk_size: int = 300, overlap: int = 50) -> List[Dict[str, Any]]:
    """
    Chunk text optimized for voice conversations.
    Smaller chunks (300 chars) for quick, focused responses.

    Args:
        text: Input text to chunk
        chunk_size: Target size for each chunk (default 300 for conversations)
        overlap: Overlap between chunks for context continuity

    Returns:
        List of chunk dictionaries with text and metadata
    """
    if not text:
        return []

    # Split by paragraphs first
    paragraphs = re.split(r'\n\s*\n', text)
    chunks = []
    chunk_id = 0

    for para_idx, paragraph in enumerate(paragraphs):
        paragraph = paragraph.strip()
        if not paragraph:
            continue

        # If paragraph is small enough, keep it as one chunk
        if len(paragraph) <= chunk_size:
            chunks.append({
                'id': f"chunk_{chunk_id}",
                'text': paragraph,
                'paragraph_id': para_idx,
                'char_count': len(paragraph)
            })
            chunk_id += 1
            continue

        # Split long paragraphs by sentences
        sentences = re.split(r'(?<=[.!?])\s+', paragraph)
        current_chunk = ""

        for sentence in sentences:
            # If adding this sentence exceeds chunk_size, save current chunk
            if len(current_chunk) + len(sentence) > chunk_size and current_chunk:
                chunks.append({
                    'id': f"chunk_{chunk_id}",
                    'text': current_chunk.strip(),
                    'paragraph_id': para_idx,
                    'char_count': len(current_chunk)
                })
                chunk_id += 1

                # Start new chunk with overlap (last sentence)
                last_sentence = current_chunk.split('.')[-2] + '.' if '.' in current_chunk else ""
                current_chunk = last_sentence + " " + sentence if last_sentence else sentence
            else:
                current_chunk += " " + sentence if current_chunk else sentence

        # Add remaining text as final chunk
        if current_chunk.strip():
            chunks.append({
                'id': f"chunk_{chunk_id}",
                'text': current_chunk.strip(),
                'paragraph_id': para_idx,
                'char_count': len(current_chunk)
            })
            chunk_id += 1

    return chunks


def create_embeddings_batch(texts: List[str], api_key: str = None) -> List[List[float]]:
    """Create embeddings using local model or OpenAI (based on EMBEDDING_PROVIDER env var)"""
    try:
        return create_embeddings_auto(texts, api_key=api_key)
    except Exception as e:
        logger.error(f"Error creating embeddings: {e}")
        return []


def process_document_for_conversation(
    assistant_id: str,
    file_path: str,
    filename: str,
    file_type: str,
    api_key: str
) -> Dict[str, Any]:
    """
    Process document and store in ChromaDB for fast conversational retrieval

    Args:
        assistant_id: AI Assistant ID (used as collection name)
        file_path: Path to the file
        filename: Original filename
        file_type: Type of file (pdf, docx, etc.)
        api_key: OpenAI API key for embeddings

    Returns:
        Dict with processing results
    """
    try:
        # Extract text
        text = extract_text_from_file(file_path, file_type)
        if not text:
            raise ValueError("Could not extract text from file")

        # Chunk text for conversations
        chunks = chunk_text_for_conversation(text, chunk_size=300, overlap=50)
        if not chunks:
            raise ValueError("No chunks created from text")

        logger.info(f"Created {len(chunks)} conversation-optimized chunks from {filename}")

        # Create embeddings
        chunk_texts = [chunk['text'] for chunk in chunks]
        embeddings = create_embeddings_batch(chunk_texts, api_key)
        if not embeddings:
            raise ValueError("Failed to create embeddings")

        # Get or create collection for this assistant
        collection_name = f"assistant_{assistant_id}"
        chroma = get_chroma_client()
        try:
            collection = chroma.get_collection(name=collection_name)
        except:
            collection = chroma.create_collection(
                name=collection_name,
                metadata={"assistant_id": assistant_id}
            )

        # Prepare data for ChromaDB
        ids = [f"{filename}_{chunk['id']}" for chunk in chunks]
        metadatas = [
            {
                'filename': filename,
                'paragraph_id': chunk['paragraph_id'],
                'char_count': chunk['char_count'],
                'file_type': file_type
            }
            for chunk in chunks
        ]

        # Add to ChromaDB
        collection.add(
            embeddings=embeddings,
            documents=chunk_texts,
            metadatas=metadatas,
            ids=ids
        )

        logger.info(f"Successfully stored {len(chunks)} chunks in ChromaDB")

        return {
            'success': True,
            'chunks_count': len(chunks),
            'text_length': len(text),
            'collection_name': collection_name
        }

    except Exception as e:
        logger.error(f"Error processing document: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'success': False,
            'error': str(e)
        }


async def search_conversation_context(
    assistant_id: str,
    query: str,
    api_key: str,
    top_k: int = 3,
    relevance_threshold: float = 0.7,
    database_config: Optional[Dict[str, Any]] = None
) -> Optional[str]:
    """
    Search knowledge base and database (if configured) and return conversational context.
    Optimized for voice conversations - returns concise, relevant info.

    Args:
        assistant_id: AI Assistant ID
        query: User's question or conversation context
        api_key: OpenAI API key
        top_k: Number of results to retrieve
        relevance_threshold: Minimum similarity score (0-1)
        database_config: Optional database configuration for querying user data

    Returns:
        Formatted context string or None if no relevant info found
    """
    context_parts = []

    # Search knowledge base (documents)
    try:
        collection_name = f"assistant_{assistant_id}"

        # Check if collection exists
        try:
            chroma = get_chroma_client()
            collection = chroma.get_collection(name=collection_name)

            # Create query embedding
            query_embeddings = create_embeddings_auto([query], api_key=api_key)
            if not query_embeddings:
                logger.warning("[CHROMA] Failed to create query embedding")
                return None
            query_embedding = query_embeddings[0]

            # Search ChromaDB
            results = collection.query(
                query_embeddings=[query_embedding],
                n_results=top_k
            )

            if results['documents'] and results['documents'][0]:
                # Build context from documents
                for i, (doc, metadata, distance) in enumerate(zip(
                    results['documents'][0],
                    results['metadatas'][0],
                    results['distances'][0]
                )):
                    # Convert distance to similarity (ChromaDB returns L2 distance)
                    similarity = 1 / (1 + distance)

                    # Only include if above threshold
                    if similarity >= relevance_threshold:
                        source = metadata.get('filename', 'document')
                        context_parts.append(f"[From {source}]: {doc}")

        except Exception as e:
            logger.info(f"No knowledge base found for assistant {assistant_id}: {e}")

    except Exception as e:
        logger.error(f"Error searching knowledge base: {e}")

    # Search database if configured
    if database_config and database_config.get('enabled'):
        try:
            from app.routes.ai_assistant.database import query_database
            from app.models.ai_assistant import DatabaseConfig

            # Convert dict to DatabaseConfig model
            db_config = DatabaseConfig(**database_config)
            db_results = await query_database(db_config, query)

            if db_results and db_results.get('records'):
                # Format database results for conversation
                db_context = "\n\n[From Database]:\n"
                for record in db_results['records'][:3]:  # Limit to top 3 records
                    # Format record fields
                    record_str = ", ".join([f"{k}: {v}" for k, v in record.items() if v is not None])
                    db_context += f"- {record_str}\n"

                context_parts.append(db_context.strip())
            elif db_results and db_results.get('documents'):
                # MongoDB results
                db_context = "\n\n[From Database]:\n"
                for doc in db_results['documents'][:3]:  # Limit to top 3 documents
                    doc_str = ", ".join([f"{k}: {v}" for k, v in doc.items() if k != '_id' and v is not None])
                    db_context += f"- {doc_str}\n"

                context_parts.append(db_context.strip())

        except Exception as e:
            logger.error(f"Error searching database: {e}")

    if not context_parts:
        return None

    # Format for conversation
    context = "Relevant information:\n" + "\n\n".join(context_parts)
    return context


def delete_document_from_kb(assistant_id: str, filename: str) -> bool:
    """
    Delete all chunks of a specific document from the knowledge base

    Args:
        assistant_id: AI Assistant ID
        filename: Name of file to delete

    Returns:
        True if successful, False otherwise
    """
    try:
        collection_name = f"assistant_{assistant_id}"

        try:
            chroma = get_chroma_client()
            collection = chroma.get_collection(name=collection_name)
        except:
            logger.warning(f"Collection not found: {collection_name}")
            return False

        # Get all IDs for this filename
        all_data = collection.get()
        ids_to_delete = [
            id for id, metadata in zip(all_data['ids'], all_data['metadatas'])
            if metadata.get('filename') == filename
        ]

        if ids_to_delete:
            collection.delete(ids=ids_to_delete)
            logger.info(f"Deleted {len(ids_to_delete)} chunks for {filename}")
            return True

        return False

    except Exception as e:
        logger.error(f"Error deleting document: {e}")
        return False


def get_kb_stats(assistant_id: str) -> Dict[str, Any]:
    """Get statistics about the knowledge base for an assistant"""
    try:
        collection_name = f"assistant_{assistant_id}"

        try:
            chroma = get_chroma_client()
            collection = chroma.get_collection(name=collection_name)
            count = collection.count()

            # Get unique files
            all_data = collection.get()
            unique_files = set(m.get('filename') for m in all_data['metadatas'])

            return {
                'exists': True,
                'total_chunks': count,
                'files_count': len(unique_files),
                'files': list(unique_files)
            }
        except:
            return {
                'exists': False,
                'total_chunks': 0,
                'files_count': 0,
                'files': []
            }
    except Exception as e:
        logger.error(f"Error getting KB stats: {e}")
        return {'exists': False, 'error': str(e)}
